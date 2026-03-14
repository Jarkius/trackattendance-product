import argparse
import itertools
import json
import random
import sys
import time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

from PyQt6.QtCore import QEventLoop, QTimer

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from attendance import AttendanceService
from sync import SyncService
import config
from main import (
    Api,
    DATABASE_PATH,
    EMPLOYEE_WORKBOOK_PATH,
    EXPORT_DIRECTORY,
    initialize_app,
)

from openpyxl import load_workbook
import requests

def _get_local_dashboard_stats(service: AttendanceService) -> Dict[str, Any]:
    """Get dashboard stats from local SQLite — same data the local dashboard shows."""
    db = service._db
    stats = db.get_sync_statistics()
    all_scans = db.fetch_all_scans()

    unique_badges = set()
    bu_counts: Dict[str, int] = {}
    station_counts: Dict[str, int] = {}

    for scan in all_scans:
        badge = scan.legacy_id
        unique_badges.add(badge)
        bu = scan.sl_l1_desc or 'Unknown'
        bu_counts[bu] = bu_counts.get(bu, 0) + 1
        station = scan.station_name or 'Unknown'
        if station not in station_counts:
            station_counts[station] = set()
        station_counts[station].add(badge)

    bu_unique: Dict[str, int] = {}
    for scan in all_scans:
        bu = scan.sl_l1_desc or 'Unknown'
        if bu not in bu_unique:
            bu_unique[bu] = set()
        bu_unique[bu].add(scan.legacy_id)

    employees_by_bu = db.get_employees_by_bu()
    bu_registered = {row['bu_name']: row['count'] for row in employees_by_bu}

    return {
        'total_scans': len(all_scans),
        'unique_badges': len(unique_badges),
        'total_registered': sum(bu_registered.values()),
        'stations': [
            {'name': name, 'unique': len(badges)}
            for name, badges in sorted(station_counts.items(), key=lambda x: -len(x[1]))
        ],
        'business_units': [
            {
                'name': bu,
                'registered': bu_registered.get(bu, 0),
                'unique': len(badges),
            }
            for bu, badges in sorted(bu_unique.items(), key=lambda x: -len(x[1]))
        ],
    }


def _get_cloud_dashboard_stats() -> Dict[str, Any]:
    """Fetch public dashboard stats from cloud API."""
    try:
        r = requests.get(
            f'{config.CLOUD_API_URL}/v1/dashboard/public/stats',
            timeout=15,
        )
        if r.status_code == 200:
            return r.json()
        return {'error': f'HTTP {r.status_code}'}
    except Exception as e:
        return {'error': str(e)}


def _compare_dashboards(local: Dict, cloud: Dict) -> List[str]:
    """Compare local vs cloud dashboard stats. Returns list of mismatches."""
    mismatches = []

    if cloud.get('error'):
        mismatches.append(f'Cloud API error: {cloud["error"]}')
        return mismatches

    if local['unique_badges'] != cloud.get('unique_badges', 0):
        mismatches.append(
            f'Unique badges: local={local["unique_badges"]} vs cloud={cloud.get("unique_badges", 0)}'
        )

    if local['total_scans'] != cloud.get('total_scans', 0):
        mismatches.append(
            f'Total scans: local={local["total_scans"]} vs cloud={cloud.get("total_scans", 0)}'
        )

    local_bus = {b['name']: b for b in local.get('business_units', [])}
    cloud_bus = {b['name']: b for b in cloud.get('business_units', [])}

    all_bu_names = set(local_bus.keys()) | set(cloud_bus.keys())
    for bu_name in sorted(all_bu_names):
        lb = local_bus.get(bu_name, {})
        cb = cloud_bus.get(bu_name, {})
        if lb.get('unique', 0) != cb.get('unique', 0):
            mismatches.append(
                f'BU "{bu_name}": local={lb.get("unique", 0)} vs cloud={cb.get("unique", 0)}'
            )

    local_stations = {s['name']: s['unique'] for s in local.get('stations', [])}
    cloud_stations = {s['name']: s['unique'] for s in cloud.get('stations', [])}

    all_station_names = set(local_stations.keys()) | set(cloud_stations.keys())
    for sn in sorted(all_station_names):
        if local_stations.get(sn, 0) != cloud_stations.get(sn, 0):
            mismatches.append(
                f'Station "{sn}": local={local_stations.get(sn, 0)} vs cloud={cloud_stations.get(sn, 0)}'
            )

    return mismatches


SPECIAL_CASE_BARCODES = [
    '999999',               # intentionally invalid control scan
    '!@#$%',                # punctuation-only input
    '12345-ABC',            # mixed digits and hyphen
    'DROP TABLE;',          # SQL-ish input
    '"quoted"',            # quotes inside payload
    "');--",               # closing quote + comment
]


def _cycle_barcodes(iterations: int, base: Iterable[str]) -> List[str]:
    base_list = list(base)
    if not base_list:
        return []
    iterator = itertools.cycle(base_list)
    return [next(iterator) for _ in range(iterations)]


def _load_employee_barcodes(workbook_path: Path) -> List[str]:
    if not workbook_path.exists():
        return []
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        header_map = {
            str(name).strip(): idx
            for idx, name in enumerate(header_row)
            if name and str(name).strip()
        }
        legacy_index = header_map.get('Legacy ID')
        if legacy_index is None:
            return []
        seen = set()
        barcodes: List[str] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cell = row[legacy_index]
            if cell is None:
                continue
            value = str(cell).strip()
            if not value or value in seen:
                continue
            seen.add(value)
            barcodes.append(value)
        return barcodes
    finally:
        workbook.close()


def _sample_employee_barcodes(workbook_path: Path, sample_size: int) -> List[str]:
    employees = _load_employee_barcodes(workbook_path)
    if not employees:
        return []
    sample_size = max(1, min(sample_size, len(employees)))
    if sample_size == len(employees):
        return employees
    return random.sample(employees, sample_size)


def _default_barcodes(sample_size: int, include_specials: bool = True) -> List[str]:
    employee_samples = _sample_employee_barcodes(EMPLOYEE_WORKBOOK_PATH, sample_size)
    barcodes = employee_samples
    if include_specials:
        barcodes = barcodes + SPECIAL_CASE_BARCODES
    if not barcodes:
        return SPECIAL_CASE_BARCODES if include_specials else []
    return barcodes


def _ensure_station_name(service: AttendanceService, fallback: str = 'Stress Harness') -> str:
    try:
        station = service.station_name
        if station:
            return station
    except RuntimeError:
        pass

    name = fallback.strip() or 'Station'
    service._db.set_station_name(name)  # type: ignore[attr-defined]
    service._station_name = name  # type: ignore[attr-defined]
    return name


def _run_js(view, script: str) -> Any:
    loop = QEventLoop()
    result_container: Dict[str, Any] = {}

    def handle_result(result: Any) -> None:
        result_container['value'] = result
        loop.quit()

    view.page().runJavaScript(script, handle_result)
    loop.exec()
    return result_container.get('value')


def _dispatch_scan(view, barcode: str) -> Dict[str, Any]:
    payload = json.dumps(barcode)
    script = f"""
        (function(barcode) {{
            const barcodeInput = document.getElementById('barcode-input');
            const feedback = document.getElementById('live-feedback-name');
            const totalScanned = document.getElementById('total-scanned');
            const historyList = document.getElementById('scan-history-list');
            if (!barcodeInput) {{
                return {{ status: 'missing-input', barcode }};
            }}
            barcodeInput.value = barcode;
            const event = new KeyboardEvent('keyup', {{ key: 'Enter', bubbles: true }});
            barcodeInput.dispatchEvent(event);
            const firstHistory = historyList && historyList.firstElementChild;
            const historyName = firstHistory ? firstHistory.querySelector('.name') : null;
            return {{
                status: 'ok',
                barcode,
                feedbackText: feedback ? feedback.textContent.trim() : null,
                feedbackColor: feedback ? window.getComputedStyle(feedback).color : null,
                totalScanned: totalScanned ? totalScanned.textContent.trim() : null,
                historyTop: historyName ? historyName.textContent.trim() : null
            }};
        }})({payload});
    """
    result = _run_js(view, script)
    if not isinstance(result, dict):
        return {'status': 'no-result', 'barcode': barcode}
    if 'barcode' not in result:
        result['barcode'] = barcode
    return result


def _collect_snapshot(view) -> Dict[str, Any]:
    script = """
        (function() {
            const feedback = document.getElementById('live-feedback-name');
            const totalScanned = document.getElementById('total-scanned');
            const historyList = document.getElementById('scan-history-list');
            const historyName = historyList && historyList.firstElementChild ? historyList.firstElementChild.querySelector('.name') : null;
            return {
                feedbackText: feedback ? feedback.textContent.trim() : null,
                feedbackColor: feedback ? window.getComputedStyle(feedback).color : null,
                totalScanned: totalScanned ? totalScanned.textContent.trim() : null,
                historyTop: historyName ? historyName.textContent.trim() : null
            };
        })();
    """
    snapshot = _run_js(view, script)
    return snapshot if isinstance(snapshot, dict) else {}


def run_stress_test(
    iterations: int,
    barcodes: Sequence[str],
    delay_ms: int,
    show_window: bool,
    show_full_screen: bool,
    enable_fade: bool,
    verbose: bool,
) -> int:
    load_state: Dict[str, Any] = {'ok': None}
    load_loop = QEventLoop()

    service = AttendanceService(
        database_path=DATABASE_PATH,
        employee_workbook_path=EMPLOYEE_WORKBOOK_PATH,
        export_directory=EXPORT_DIRECTORY,
    )
    _ensure_station_name(service)

    # Initialize sync service for cloud integration testing
    sync_service = SyncService(
        db=service._db,
        api_url=config.CLOUD_API_URL,
        api_key=config.CLOUD_API_KEY,
        batch_size=config.CLOUD_SYNC_BATCH_SIZE,
    )

    def on_load_finished(ok: bool) -> None:
        load_state['ok'] = ok
        if load_loop.isRunning():
            load_loop.quit()

    def api_factory(quit_callback):
        return Api(service=service, quit_callback=quit_callback, sync_service=sync_service)

    try:
        app, window, view, _ = initialize_app(
            argv=sys.argv,
            show_window=show_window,
            show_full_screen=show_full_screen,
            enable_fade=enable_fade,
            on_load_finished=on_load_finished,
            api_factory=api_factory,
        )

        window.setProperty('suppress_export_notification', True)

        # Maximize window if showing in windowed mode
        if show_window and not show_full_screen:
            window.showMaximized()

        if load_state['ok'] is None:
            load_loop.exec()
        if not load_state['ok']:
            return 3

        sequence = _cycle_barcodes(iterations, barcodes)
        failures: List[Dict[str, Any]] = []
        successful = 0
        invalid = 0
        start = time.perf_counter()

        for index, barcode in enumerate(sequence, start=1):
            result = _dispatch_scan(view, barcode)
            status = result.get('status', 'unknown')
            feedback = (result.get('feedbackText') or '').strip()
            total_scanned = result.get('totalScanned')
            history_top = result.get('historyTop')

            if status != 'ok':
                failures.append(result)
            else:
                if feedback.lower().startswith('not matched'):
                    invalid += 1
                else:
                    successful += 1

            if verbose:
                print(f"[{index:03d}] status={status} barcode={barcode} feedback='{feedback}' total={total_scanned} history='{history_top}'")
            elif index % 25 == 0 or index == 1 or index == iterations:
                print(f"[{index:03d}/{iterations}] status={status} feedback='{feedback}' total={total_scanned}")

            # Update UI sync statistics every 100 scans to show pending count growing
            if index % 100 == 0 or index == iterations:
                view.page().runJavaScript("""
                    (function() {
                        if (typeof updateSyncStatus === 'function') {
                            updateSyncStatus();
                        }
                    })();
                """)

            if delay_ms > 0:
                settle_loop = QEventLoop()
                QTimer.singleShot(delay_ms, settle_loop.quit)
                settle_loop.exec()

        duration = time.perf_counter() - start
        snapshot = _collect_snapshot(view)

        # Test sync service before export (simulates shutdown sync)
        sync_attempted = False
        sync_success = False
        synced_count = 0
        failed_count = 0

        if sync_service:
            try:
                stats_before = sync_service.db.get_sync_statistics()
                pending_before = stats_before.get('pending', 0)

                if pending_before > 0:
                    sync_attempted = True
                    print(f'[sync] {pending_before} pending scan(s) detected. Waiting 2 seconds to view UI...')

                    # Update UI to show pending count before sync
                    view.page().runJavaScript("""
                        (function() {
                            if (typeof updateSyncStatus === 'function') {
                                updateSyncStatus();
                            }
                        })();
                    """)
                    time.sleep(2)  # Give user time to see pending count in UI

                    # Check connection before attempting sync
                    print(f'[sync] Testing connection to cloud API...')
                    connection_ok, connection_msg = sync_service.test_connection()

                    if not connection_ok:
                        print(f'[sync] [X] No connection to cloud API: {connection_msg}')
                        print(f'[sync] Skipping sync - {pending_before} scan(s) will remain pending')
                        sync_attempted = False
                        sync_success = False
                    else:
                        print(f'[sync] [OK] Connected to cloud API')
                        print(f'[sync] Syncing {pending_before} pending scan(s) now...')
                        sync_start = time.perf_counter()

                        # Sync all pending scans using sync_all=True
                        sync_result = sync_service.sync_pending_scans(sync_all=True)
                        sync_duration = time.perf_counter() - sync_start

                        synced_count = sync_result.get('synced', 0)
                        failed_count = sync_result.get('failed', 0)
                        pending_after = sync_result.get('pending', 0)
                        batch_count = sync_result.get('batches', 0)

                        sync_success = synced_count > 0 or failed_count == 0

                        print(f'[sync] Complete in {sync_duration:.2f}s: {synced_count} synced, {failed_count} failed, {pending_after} pending ({batch_count} batches)')

                        # Update UI sync statistics after sync completes
                        view.page().runJavaScript("""
                            (function() {
                                if (typeof updateSyncStatus === 'function') {
                                    updateSyncStatus();
                                }
                            })();
                        """)

                        # Keep window open for 3 seconds after sync completes so user can see results
                        print(f'[sync] Keeping window open for 3 seconds to view results...')
                        time.sleep(3)
                else:
                    print('[sync] No pending scans to sync')
            except Exception as exc:
                print(f'[sync] Sync failed: {exc}')
                sync_attempted = True
                sync_success = False

        # ---- Dashboard Comparison: Local vs Cloud ----
        dashboard_match = None
        if sync_attempted and sync_success:
            print(f'\n[dashboard] Comparing local vs cloud dashboard data...')
            try:
                # Sync roster summary so cloud has registered counts
                from sync import sync_roster_summary_from_data
                bu_data = service._db.get_employees_by_bu()
                if bu_data:
                    sync_roster_summary_from_data(bu_data, config.CLOUD_API_URL, config.CLOUD_API_KEY)

                local_stats = _get_local_dashboard_stats(service)
                cloud_stats = _get_cloud_dashboard_stats()

                print(f'\n{"="*60}')
                print(f'LOCAL DASHBOARD')
                print(f'{"="*60}')
                print(f'Total registered: {local_stats["total_registered"]}')
                print(f'Unique badges:    {local_stats["unique_badges"]}')
                print(f'Total scans:      {local_stats["total_scans"]}')
                rate = (local_stats["unique_badges"] / local_stats["total_registered"] * 100) if local_stats["total_registered"] > 0 else 0
                print(f'Attendance rate:  {rate:.1f}%')
                for s in local_stats['stations']:
                    print(f'  Station {s["name"]}: {s["unique"]} unique')
                for bu in local_stats['business_units']:
                    pct = f' ({bu["unique"]*100/bu["registered"]:.1f}%)' if bu['registered'] > 0 else ''
                    print(f'  BU {bu["name"]}: {bu["unique"]}/{bu["registered"]}{pct}')

                print(f'\n{"="*60}')
                print(f'CLOUD DASHBOARD (mobile)')
                print(f'{"="*60}')
                if cloud_stats.get('error'):
                    print(f'  [ERROR] {cloud_stats["error"]}')
                else:
                    print(f'Total registered: {cloud_stats.get("total_registered", "N/A")}')
                    print(f'Unique badges:    {cloud_stats.get("unique_badges", 0)}')
                    print(f'Total scans:      {cloud_stats.get("total_scans", 0)}')
                    for s in cloud_stats.get('stations', []):
                        print(f'  Station {s["name"]}: {s["unique"]} unique')
                    for bu in cloud_stats.get('business_units', []):
                        pct = f' ({bu["unique"]*100/bu["registered"]:.1f}%)' if bu.get('registered', 0) > 0 else ''
                        print(f'  BU {bu["name"]}: {bu["unique"]}/{bu.get("registered", 0)}{pct}')

                mismatches = _compare_dashboards(local_stats, cloud_stats)
                print(f'\n{"="*60}')
                if mismatches:
                    dashboard_match = False
                    print(f'DASHBOARD COMPARISON: MISMATCH ({len(mismatches)} differences)')
                    for m in mismatches:
                        print(f'  [X] {m}')
                else:
                    dashboard_match = True
                    print(f'DASHBOARD COMPARISON: MATCH — local and cloud data are identical')
                print(f'{"="*60}')
            except Exception as exc:
                print(f'[dashboard] Comparison failed: {exc}')

        export_info = None
        try:
            export_info = service.export_scans()
        except Exception as exc:
            if verbose:
                print(f'[warn] export failed: {exc}')
        else:
            if export_info and verbose:
                dest = export_info.get('absolutePath') or export_info.get('fileName')
                if dest:
                    print(f'[info] export written to {dest}')
            window.setProperty('export_notification_triggered', True)

        # Final UI update to ensure user sees pending=0 and updated sync counters
        # Uses direct DOM manipulation (Issue #11) to avoid async callback timing issues
        if sync_attempted and sync_success:
            print(f'[info] Updating UI to show final sync status (pending=0)...')

            # Verify actual database state first
            final_stats = sync_service.db.get_sync_statistics()
            print(f'[info] Database stats - Pending: {final_stats["pending"]}, Synced: {final_stats["synced"]}, Failed: {final_stats["failed"]}')

            # Option A: Direct DOM Manipulation (Synchronous, No Callbacks)
            # This bypasses the async callback chain that causes timing issues
            print(f'[info] Updating UI via direct DOM manipulation...')
            view.page().runJavaScript(f"""
                (function() {{
                    // Directly set DOM values from sync results
                    var pending = document.getElementById('sync-pending');
                    var synced = document.getElementById('sync-synced');
                    var failed = document.getElementById('sync-failed');

                    if (pending) {{
                        pending.textContent = '{final_stats["pending"]}';
                        console.log('[UI Update] Set pending=' + pending.textContent);
                    }}
                    if (synced) {{
                        synced.textContent = '{final_stats["synced"]}';
                        console.log('[UI Update] Set synced=' + synced.textContent);
                    }}
                    if (failed) {{
                        failed.textContent = '{final_stats["failed"]}';
                        console.log('[UI Update] Set failed=' + failed.textContent);
                    }}

                    // Log to console for verification
                    console.log('[UI Update] Direct DOM update complete - pending={final_stats["pending"]}, synced={final_stats["synced"]}, failed={final_stats["failed"]}');
                }})();
            """)

            # Process events to render the changes
            print(f'[info] Processing events to render DOM changes...')
            for _ in range(20):
                app.processEvents()

            # Wait for visual rendering to complete
            print(f'[info] Waiting for visual rendering...')
            time.sleep(2)

            # Now keep window open so user can see the results
            print(f'[info] ============================================')
            print(f'[info] Window will stay open for 10 seconds')
            print(f'[info] VERIFY IN UI: Pending: {final_stats["pending"]}, Synced: {final_stats["synced"]}, Failed: {final_stats["failed"]}')
            print(f'[info] ============================================')

            # Sleep in small chunks while processing events to keep UI responsive
            for i in range(100):  # 100 × 100ms = 10 seconds
                time.sleep(0.1)
                app.processEvents()

        window.close()
        for _ in range(3):
            app.processEvents()

        print("\n--- Stress Test Summary ---")
        print(f'Scans attempted : {iterations}')
        print(f'Successful scans: {successful}')
        print(f'Invalid scans   : {invalid}')
        print(f'Failures        : {len(failures)}')
        print(f'Total runtime   : {duration:.2f}s')

        if sync_attempted:
            print(f'\n--- Cloud Sync Results ---')
            print(f'Sync attempted  : Yes')
            print(f'Sync success    : {"Yes" if sync_success else "No"}')
            print(f'Scans synced    : {synced_count}')
            print(f'Scans failed    : {failed_count}')
            if dashboard_match is not None:
                print(f'Dashboard match : {"Yes" if dashboard_match else "MISMATCH"}')
        else:
            print(f'\n--- Cloud Sync Results ---')
            print(f'Sync attempted  : No (no pending scans)')

        if failures:
            print("\nFirst failure:")
            print(json.dumps(failures[0], indent=2))
            return 2

        expected_total = str(successful)
        final_total = snapshot.get('totalScanned') if snapshot else None
        if final_total and final_total != expected_total:
            print(f"Warning: total_scanned reported as {final_total}, expected {expected_total}.")

        return 0
    finally:
        service.close()


def main() -> int:
    parser = argparse.ArgumentParser(description='Drive the full PyQt window and simulate barcode scans.')
    parser.add_argument('barcodes', nargs='*', help='Base barcode values to cycle through; defaults use random workbook samples.')
    parser.add_argument('--iterations', type=int, default=200, help='Number of barcode submissions to perform.')
    parser.add_argument('--delay-ms', type=int, default=75, help='Delay between scans to mimic hardware pacing.')
    parser.add_argument('--sample-size', type=int, default=50, help='Employee barcodes to sample from the workbook when no explicit list is provided.')
    parser.add_argument('--no-specials', action='store_true', help='Exclude synthetic invalid barcode cases from the run.')
    parser.add_argument('--no-show-window', action='store_true', help='Keep the window hidden during the run.')
    parser.add_argument('--fullscreen', action='store_true', help='Show the window in fullscreen mode (default is maximized window).')
    parser.add_argument('--disable-fade', action='store_true', help='Skip the window fade animation to save a few frames.')
    parser.add_argument('--verbose', action='store_true', help='Log every scan instead of periodic checkpoints.')
    args = parser.parse_args()

    if args.barcodes:
        base_barcodes = list(args.barcodes)
    else:
        base_barcodes = _default_barcodes(sample_size=max(args.sample_size, 1), include_specials=not args.no_specials)

    if not base_barcodes:
        print('No barcodes available for the stress test.', file=sys.stderr)
        return 1

    status = run_stress_test(
        iterations=args.iterations,
        barcodes=base_barcodes,
        delay_ms=max(args.delay_ms, 0),
        show_window=not args.no_show_window,
        show_full_screen=args.fullscreen,
        enable_fade=not args.disable_fade,
        verbose=args.verbose,
    )
    return status


if __name__ == '__main__':
    sys.exit(main())
