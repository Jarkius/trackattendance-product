#!/usr/bin/env python3
"""
Tests for duplicate badge detection functionality.

Tests the duplicate detection in database.py and attendance.py including:
- Database layer: check_if_duplicate_badge()
- All three detection modes: block, warn, silent
- Time window boundary conditions
- Configuration options

Run: python tests/test_duplicate_detection.py
"""

import os
import sys
import tempfile
import time
import unittest
from pathlib import Path
from datetime import datetime, timezone, timedelta
from unittest.mock import patch

# Set required environment variables BEFORE importing config
os.environ.setdefault("CLOUD_API_KEY", "test-api-key-for-testing")
os.environ.setdefault("CLOUD_API_URL", "http://test.example.com")

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from database import DatabaseManager, EmployeeRecord, ISO_TIMESTAMP_FORMAT


# Check if PyQt6 is available for AttendanceService tests
try:
    from PyQt6.QtWidgets import QApplication
    PYQT6_AVAILABLE = True
except ImportError:
    PYQT6_AVAILABLE = False


def _create_service(db_path, employee_path, export_dir, station_name="TestStation"):
    """Create an AttendanceService with station name pre-configured for testing."""
    from attendance import AttendanceService
    service = AttendanceService(
        database_path=db_path,
        employee_workbook_path=employee_path,
        export_directory=export_dir,
    )
    service._db.set_station_name(station_name)
    service._station_name = station_name
    return service


class TestDatabaseDuplicateDetection(unittest.TestCase):
    """Test database layer duplicate detection (check_if_duplicate_badge)."""

    def setUp(self):
        """Set up test database."""
        self.temp_dir = tempfile.mkdtemp()
        self.db_path = Path(self.temp_dir) / "test.db"
        self.db = DatabaseManager(self.db_path)
        self.db.set_station_name("TestStation")

        # Add test employee
        self.db.bulk_insert_employees([
            EmployeeRecord("TEST001", "Test User", "IT", "Engineer"),
            EmployeeRecord("TEST002", "Another User", "HR", "Manager"),
        ])

    def tearDown(self):
        """Clean up test resources."""
        self.db.close()
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_duplicate_within_time_window(self):
        """Test duplicate detected when same badge scanned within window."""
        employee = EmployeeRecord("TEST001", "Test User", "IT", "Engineer")

        # First scan
        self.db.record_scan("TEST001", "TestStation", employee)

        # Check for duplicate immediately
        is_duplicate, original_id = self.db.check_if_duplicate_badge(
            "TEST001", "TestStation", time_window_seconds=60
        )

        self.assertTrue(is_duplicate)
        self.assertIsNotNone(original_id)

    def test_no_duplicate_outside_time_window(self):
        """Test no duplicate when same badge scanned outside window."""
        employee = EmployeeRecord("TEST001", "Test User", "IT", "Engineer")

        # Record scan with timestamp from 2 minutes ago
        old_time = datetime.now(timezone.utc) - timedelta(seconds=120)
        old_timestamp = old_time.strftime(ISO_TIMESTAMP_FORMAT)
        self.db.record_scan("TEST001", "TestStation", employee, old_timestamp)

        # Check for duplicate with 60 second window
        is_duplicate, original_id = self.db.check_if_duplicate_badge(
            "TEST001", "TestStation", time_window_seconds=60
        )

        self.assertFalse(is_duplicate)
        self.assertIsNone(original_id)

    def test_no_duplicate_different_badge(self):
        """Test no duplicate for different badge IDs."""
        employee1 = EmployeeRecord("TEST001", "Test User", "IT", "Engineer")

        # First scan with TEST001
        self.db.record_scan("TEST001", "TestStation", employee1)

        # Check duplicate for different badge
        is_duplicate, original_id = self.db.check_if_duplicate_badge(
            "TEST002", "TestStation", time_window_seconds=60
        )

        self.assertFalse(is_duplicate)
        self.assertIsNone(original_id)

    def test_duplicate_same_station_only(self):
        """Test duplicate detection is station-specific."""
        employee = EmployeeRecord("TEST001", "Test User", "IT", "Engineer")

        # Scan at Station A
        self.db.record_scan("TEST001", "StationA", employee)

        # Check duplicate at Station B (should NOT be duplicate)
        is_duplicate, original_id = self.db.check_if_duplicate_badge(
            "TEST001", "StationB", time_window_seconds=60
        )

        self.assertFalse(is_duplicate)

        # Check duplicate at Station A (should BE duplicate)
        is_duplicate, original_id = self.db.check_if_duplicate_badge(
            "TEST001", "StationA", time_window_seconds=60
        )

        self.assertTrue(is_duplicate)

    def test_time_window_boundary_inside(self):
        """Test scan exactly at window edge (inside) is duplicate."""
        employee = EmployeeRecord("TEST001", "Test User", "IT", "Engineer")

        # Record scan exactly 59 seconds ago (inside 60s window)
        past_time = datetime.now(timezone.utc) - timedelta(seconds=59)
        past_timestamp = past_time.strftime(ISO_TIMESTAMP_FORMAT)
        self.db.record_scan("TEST001", "TestStation", employee, past_timestamp)

        # Should be duplicate (59 < 60)
        is_duplicate, _ = self.db.check_if_duplicate_badge(
            "TEST001", "TestStation", time_window_seconds=60
        )

        self.assertTrue(is_duplicate)

    def test_time_window_boundary_outside(self):
        """Test scan just outside window is not duplicate."""
        employee = EmployeeRecord("TEST001", "Test User", "IT", "Engineer")

        # Record scan exactly 61 seconds ago (outside 60s window)
        past_time = datetime.now(timezone.utc) - timedelta(seconds=61)
        past_timestamp = past_time.strftime(ISO_TIMESTAMP_FORMAT)
        self.db.record_scan("TEST001", "TestStation", employee, past_timestamp)

        # Should NOT be duplicate (61 > 60)
        is_duplicate, _ = self.db.check_if_duplicate_badge(
            "TEST001", "TestStation", time_window_seconds=60
        )

        self.assertFalse(is_duplicate)

    def test_empty_database_no_duplicate(self):
        """Test no duplicate in empty database."""
        is_duplicate, original_id = self.db.check_if_duplicate_badge(
            "NONEXISTENT", "TestStation", time_window_seconds=60
        )

        self.assertFalse(is_duplicate)
        self.assertIsNone(original_id)

    def test_custom_time_window(self):
        """Test custom time window values."""
        employee = EmployeeRecord("TEST001", "Test User", "IT", "Engineer")

        # Record scan 45 seconds ago
        past_time = datetime.now(timezone.utc) - timedelta(seconds=45)
        past_timestamp = past_time.strftime(ISO_TIMESTAMP_FORMAT)
        self.db.record_scan("TEST001", "TestStation", employee, past_timestamp)

        # With 30 second window, should NOT be duplicate
        is_duplicate, _ = self.db.check_if_duplicate_badge(
            "TEST001", "TestStation", time_window_seconds=30
        )
        self.assertFalse(is_duplicate)

        # With 60 second window, SHOULD be duplicate
        is_duplicate, _ = self.db.check_if_duplicate_badge(
            "TEST001", "TestStation", time_window_seconds=60
        )
        self.assertTrue(is_duplicate)

    def test_multiple_scans_returns_most_recent(self):
        """Test that duplicate check finds most recent scan."""
        employee = EmployeeRecord("TEST001", "Test User", "IT", "Engineer")

        # Record multiple scans
        for i in range(3):
            time_offset = 50 - (i * 10)  # 50s, 40s, 30s ago
            past_time = datetime.now(timezone.utc) - timedelta(seconds=time_offset)
            past_timestamp = past_time.strftime(ISO_TIMESTAMP_FORMAT)
            self.db.record_scan("TEST001", "TestStation", employee, past_timestamp)

        # Check should find the most recent (30s ago)
        is_duplicate, original_id = self.db.check_if_duplicate_badge(
            "TEST001", "TestStation", time_window_seconds=60
        )

        self.assertTrue(is_duplicate)
        self.assertIsNotNone(original_id)


# Check if PyQt6 is available for integration tests
try:
    from PyQt6.QtWidgets import QApplication
    PYQT6_AVAILABLE = True
except ImportError:
    PYQT6_AVAILABLE = False


@unittest.skipUnless(PYQT6_AVAILABLE, "PyQt6 not available")
class TestDuplicateDetectionBlockMode(unittest.TestCase):
    """Test 'block' mode: duplicate scans are rejected."""

    def setUp(self):
        """Set up test database and attendance service."""
        self.temp_dir = tempfile.mkdtemp()
        self.db_path = Path(self.temp_dir) / "test.db"
        self.employee_path = Path(self.temp_dir) / "employee.xlsx"
        self.export_dir = Path(self.temp_dir) / "exports"
        self.export_dir.mkdir()

        # Create employee file
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Legacy ID", "Full Name", "SL L1 Desc", "Position Desc"])
        ws.append(["TEST001", "Test User", "IT", "Engineer"])
        wb.save(self.employee_path)

    def tearDown(self):
        """Clean up."""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    @patch.dict(os.environ, {
        'DUPLICATE_BADGE_DETECTION_ENABLED': 'True',
        'DUPLICATE_BADGE_ACTION': 'block',
        'DUPLICATE_BADGE_TIME_WINDOW_SECONDS': '60',
    })
    def test_block_mode_rejects_duplicate(self):
        """Test block mode rejects duplicate scan."""
        import importlib
        import config
        importlib.reload(config)

        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            # First scan succeeds
            result1 = service.register_scan("TEST001")
            self.assertTrue(result1["ok"])
            self.assertFalse(result1.get("is_duplicate", False))

            # Second scan should be blocked
            result2 = service.register_scan("TEST001")
            self.assertFalse(result2["ok"])
            self.assertEqual(result2.get("status"), "duplicate_rejected")
            self.assertTrue(result2.get("is_duplicate", False))
        finally:
            service.close()

    @patch.dict(os.environ, {
        'DUPLICATE_BADGE_DETECTION_ENABLED': 'True',
        'DUPLICATE_BADGE_ACTION': 'block',
        'DUPLICATE_BADGE_TIME_WINDOW_SECONDS': '60',
    })
    def test_block_mode_preserves_original(self):
        """Test block mode preserves original scan in database."""
        import importlib
        import config
        importlib.reload(config)

        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            # First scan
            service.register_scan("TEST001")

            # Second scan (blocked)
            service.register_scan("TEST001")

            # Should only have 1 scan in database
            scans = service._db.fetch_all_scans()
            self.assertEqual(len(scans), 1)
        finally:
            service.close()


@unittest.skipUnless(PYQT6_AVAILABLE, "PyQt6 not available")
class TestDuplicateDetectionWarnMode(unittest.TestCase):
    """Test 'warn' mode: duplicate scans are recorded with warning."""

    def setUp(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.db_path = Path(self.temp_dir) / "test.db"
        self.employee_path = Path(self.temp_dir) / "employee.xlsx"
        self.export_dir = Path(self.temp_dir) / "exports"
        self.export_dir.mkdir()

        # Create employee file
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Legacy ID", "Full Name", "SL L1 Desc", "Position Desc"])
        ws.append(["TEST001", "Test User", "IT", "Engineer"])
        wb.save(self.employee_path)

    def tearDown(self):
        """Clean up."""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    @patch.dict(os.environ, {
        'DUPLICATE_BADGE_DETECTION_ENABLED': 'True',
        'DUPLICATE_BADGE_ACTION': 'warn',
        'DUPLICATE_BADGE_TIME_WINDOW_SECONDS': '60',
    })
    def test_warn_mode_records_duplicate(self):
        """Test warn mode records duplicate scan."""
        import importlib
        import config
        importlib.reload(config)

        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            # First scan
            result1 = service.register_scan("TEST001")
            self.assertTrue(result1["ok"])

            # Second scan (should be recorded with warning)
            result2 = service.register_scan("TEST001")
            self.assertTrue(result2["ok"])
            self.assertTrue(result2.get("is_duplicate", False))

            # Should have 2 scans in database
            scans = service._db.fetch_all_scans()
            self.assertEqual(len(scans), 2)
        finally:
            service.close()

    @patch.dict(os.environ, {
        'DUPLICATE_BADGE_DETECTION_ENABLED': 'True',
        'DUPLICATE_BADGE_ACTION': 'warn',
        'DUPLICATE_BADGE_TIME_WINDOW_SECONDS': '60',
    })
    def test_warn_mode_returns_warning_flag(self):
        """Test warn mode sets is_duplicate flag for UI."""
        import importlib
        import config
        importlib.reload(config)

        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            service.register_scan("TEST001")
            result = service.register_scan("TEST001")

            # Should have warning flag for UI (yellow alert)
            self.assertTrue(result.get("is_duplicate", False))
        finally:
            service.close()


@unittest.skipUnless(PYQT6_AVAILABLE, "PyQt6 not available")
class TestDuplicateDetectionSilentMode(unittest.TestCase):
    """Test 'silent' mode: duplicate scans are recorded without alert."""

    def setUp(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.db_path = Path(self.temp_dir) / "test.db"
        self.employee_path = Path(self.temp_dir) / "employee.xlsx"
        self.export_dir = Path(self.temp_dir) / "exports"
        self.export_dir.mkdir()

        # Create employee file
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Legacy ID", "Full Name", "SL L1 Desc", "Position Desc"])
        ws.append(["TEST001", "Test User", "IT", "Engineer"])
        wb.save(self.employee_path)

    def tearDown(self):
        """Clean up."""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    @patch.dict(os.environ, {
        'DUPLICATE_BADGE_DETECTION_ENABLED': 'True',
        'DUPLICATE_BADGE_ACTION': 'silent',
        'DUPLICATE_BADGE_TIME_WINDOW_SECONDS': '60',
    })
    def test_silent_mode_records_duplicate(self):
        """Test silent mode records duplicate without UI alert flag."""
        import importlib
        import config
        importlib.reload(config)

        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            # First scan
            service.register_scan("TEST001")

            # Second scan (recorded silently)
            result = service.register_scan("TEST001")

            # Should be OK without duplicate alert
            self.assertTrue(result["ok"])
            # is_duplicate should be False in silent mode (no UI alert)
            self.assertFalse(result.get("is_duplicate", False))

            # Should have 2 scans in database
            scans = service._db.fetch_all_scans()
            self.assertEqual(len(scans), 2)
        finally:
            service.close()


@unittest.skipUnless(PYQT6_AVAILABLE, "PyQt6 not available")
class TestDuplicateDetectionDisabled(unittest.TestCase):
    """Test duplicate detection when disabled."""

    def setUp(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.db_path = Path(self.temp_dir) / "test.db"
        self.employee_path = Path(self.temp_dir) / "employee.xlsx"
        self.export_dir = Path(self.temp_dir) / "exports"
        self.export_dir.mkdir()

        # Create employee file
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Legacy ID", "Full Name", "SL L1 Desc", "Position Desc"])
        ws.append(["TEST001", "Test User", "IT", "Engineer"])
        wb.save(self.employee_path)

    def tearDown(self):
        """Clean up."""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    @patch.dict(os.environ, {
        'DUPLICATE_BADGE_DETECTION_ENABLED': 'False',
    })
    def test_disabled_allows_all_scans(self):
        """Test disabled detection allows all scans without warnings."""
        import importlib
        import config
        importlib.reload(config)

        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            # Multiple rapid scans should all succeed
            for i in range(3):
                result = service.register_scan("TEST001")
                self.assertTrue(result["ok"])
                self.assertFalse(result.get("is_duplicate", False))

            # All scans recorded
            scans = service._db.fetch_all_scans()
            self.assertEqual(len(scans), 3)
        finally:
            service.close()


class TestDuplicateDetectionConfigVariations(unittest.TestCase):
    """Test various configuration combinations."""

    def setUp(self):
        """Set up test database."""
        self.temp_dir = tempfile.mkdtemp()
        self.db_path = Path(self.temp_dir) / "test.db"
        self.db = DatabaseManager(self.db_path)
        self.db.set_station_name("TestStation")

        self.db.bulk_insert_employees([
            EmployeeRecord("TEST001", "Test User", "IT", "Engineer")
        ])

    def tearDown(self):
        """Clean up."""
        self.db.close()
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_very_short_time_window(self):
        """Test with very short time window (1 second)."""
        employee = EmployeeRecord("TEST001", "Test User", "IT", "Engineer")

        # Record scan with current timestamp
        self.db.record_scan("TEST001", "TestStation", employee)

        # Immediate check with 1 second window should be duplicate
        is_duplicate, _ = self.db.check_if_duplicate_badge(
            "TEST001", "TestStation", time_window_seconds=1
        )
        self.assertTrue(is_duplicate)

        # Record scan 2 seconds ago (using explicit timestamp)
        past_time = datetime.now(timezone.utc) - timedelta(seconds=2)
        past_timestamp = past_time.strftime(ISO_TIMESTAMP_FORMAT)

        # Clear and re-record with old timestamp
        self.db._connection.execute("DELETE FROM scans")
        self.db.record_scan("TEST001", "TestStation", employee, past_timestamp)

        # Check with 1 second window - should NOT be duplicate (2s > 1s)
        is_duplicate, _ = self.db.check_if_duplicate_badge(
            "TEST001", "TestStation", time_window_seconds=1
        )
        self.assertFalse(is_duplicate)

    def test_very_long_time_window(self):
        """Test with long time window (1 hour)."""
        employee = EmployeeRecord("TEST001", "Test User", "IT", "Engineer")

        # Record scan 30 minutes ago
        past_time = datetime.now(timezone.utc) - timedelta(minutes=30)
        past_timestamp = past_time.strftime(ISO_TIMESTAMP_FORMAT)
        self.db.record_scan("TEST001", "TestStation", employee, past_timestamp)

        # With 1 hour window, should still be duplicate
        is_duplicate, _ = self.db.check_if_duplicate_badge(
            "TEST001", "TestStation", time_window_seconds=3600
        )
        self.assertTrue(is_duplicate)


def main():
    """Run tests with summary."""
    print("=" * 70)
    print("DUPLICATE BADGE DETECTION TESTS")
    print("=" * 70)
    print()

    # Create test suite
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()

    suite.addTests(loader.loadTestsFromTestCase(TestDatabaseDuplicateDetection))
    suite.addTests(loader.loadTestsFromTestCase(TestDuplicateDetectionBlockMode))
    suite.addTests(loader.loadTestsFromTestCase(TestDuplicateDetectionWarnMode))
    suite.addTests(loader.loadTestsFromTestCase(TestDuplicateDetectionSilentMode))
    suite.addTests(loader.loadTestsFromTestCase(TestDuplicateDetectionDisabled))
    suite.addTests(loader.loadTestsFromTestCase(TestDuplicateDetectionConfigVariations))

    # Run with verbosity
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)

    print()
    print("=" * 70)
    print(f"Tests run: {result.testsRun}")
    print(f"Failures: {len(result.failures)}")
    print(f"Errors: {len(result.errors)}")
    print("=" * 70)

    return 0 if result.wasSuccessful() else 1


if __name__ == "__main__":
    sys.exit(main())
