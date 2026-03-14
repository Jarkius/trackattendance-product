#!/usr/bin/env python3
"""
Tests for Excel export functionality in attendance.py.

Tests the export_scans() method including:
- Basic export functionality
- Thai character handling
- Special characters in data
- Edge cases (empty data, large datasets)
- Error handling

Run: python tests/test_excel_export.py
"""

import os
import sys
import tempfile
import unittest
from pathlib import Path
from datetime import datetime, timezone
from unittest.mock import patch, Mock

# Set required environment variables BEFORE importing config
os.environ.setdefault("CLOUD_API_KEY", "test-api-key-for-testing")
os.environ.setdefault("CLOUD_API_URL", "http://test.example.com")

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook, load_workbook
from database import DatabaseManager, EmployeeRecord, ISO_TIMESTAMP_FORMAT

# Check if PyQt6 is available for integration tests
try:
    from PyQt6.QtWidgets import QApplication
    PYQT6_AVAILABLE = True
except ImportError:
    PYQT6_AVAILABLE = False


def _create_service(db_path, employee_path, export_dir, station_name="TestStation"):
    """Create an AttendanceService with station name pre-configured for testing."""
    from attendance import AttendanceService
    service = AttendanceService(
        database_path=db_path,
        employee_workbook_path=employee_path,
        export_directory=export_dir,
    )
    service._db.set_station_name(station_name)
    service._station_name = station_name
    return service


@unittest.skipUnless(PYQT6_AVAILABLE, "PyQt6 not available")
class TestExcelExportBasic(unittest.TestCase):
    """Test basic Excel export functionality."""

    def setUp(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.db_path = Path(self.temp_dir) / "test.db"
        self.employee_path = Path(self.temp_dir) / "employee.xlsx"
        self.export_dir = Path(self.temp_dir) / "exports"
        self.export_dir.mkdir()

        # Create employee file
        wb = Workbook()
        ws = wb.active
        ws.append(["Legacy ID", "Full Name", "SL L1 Desc", "Position Desc"])
        ws.append(["TEST001", "Test User", "IT", "Engineer"])
        ws.append(["TEST002", "Jane Doe", "HR", "Manager"])
        wb.save(self.employee_path)

    def tearDown(self):
        """Clean up."""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_export_creates_valid_xlsx(self):
        """Test export creates valid .xlsx file."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            # Add a scan
            service.register_scan("TEST001")

            # Export
            result = service.export_scans()

            self.assertTrue(result["ok"])
            self.assertIn("fileName", result)
            self.assertTrue(result["fileName"].endswith(".xlsx"))

            # Verify file exists and is valid Excel
            export_path = Path(result["absolutePath"])
            self.assertTrue(export_path.exists())

            # Load and verify
            wb = load_workbook(export_path)
            self.assertIsNotNone(wb.active)
            wb.close()
        finally:
            service.close()

    def test_export_contains_correct_headers(self):
        """Test exported file has correct headers."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            service.register_scan("TEST001")
            result = service.export_scans()

            export_path = Path(result["absolutePath"])
            wb = load_workbook(export_path)
            ws = wb.active

            # Get headers from first row
            headers = [cell.value for cell in ws[1]]

            # Expected headers
            self.assertIn("Scan Value", headers)
            self.assertIn("Matched", headers)
            self.assertIn("Station", headers)
            self.assertIn("Scanned At", headers)
            self.assertIn("Scan Source", headers)

            wb.close()
        finally:
            service.close()

    def test_export_data_matches_database(self):
        """Test exported data matches database content."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            # Add multiple scans
            service.register_scan("TEST001")
            service.register_scan("TEST002")
            service.register_scan("UNKNOWN999")  # Unmatched

            result = service.export_scans()

            self.assertEqual(result["records"], 3)

            export_path = Path(result["absolutePath"])
            wb = load_workbook(export_path)
            ws = wb.active

            # Count data rows (excluding header)
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            self.assertEqual(len(data_rows), 3)

            # Verify badge IDs are in export
            badge_ids = [row[0] for row in data_rows]
            self.assertIn("TEST001", badge_ids)
            self.assertIn("TEST002", badge_ids)
            self.assertIn("UNKNOWN999", badge_ids)

            wb.close()
        finally:
            service.close()

    def test_export_filename_format(self):
        """Test export filename contains station and timestamp."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            # Set station name
            service._db.set_station_name("MainEntrance")
            service._station_name = "MainEntrance"

            service.register_scan("TEST001")
            result = service.export_scans()

            filename = result["fileName"]

            # Should contain "Checkins"
            self.assertIn("Checkins", filename)
            # Should contain station name (sanitized)
            self.assertIn("MainEntrance", filename)
            # Should end with .xlsx
            self.assertTrue(filename.endswith(".xlsx"))

            wb = load_workbook(result["absolutePath"])
            wb.close()
        finally:
            service.close()


@unittest.skipUnless(PYQT6_AVAILABLE, "PyQt6 not available")
class TestExcelExportThaiCharacters(unittest.TestCase):
    """Test Thai character handling in exports."""

    def setUp(self):
        """Set up test environment with Thai data."""
        self.temp_dir = tempfile.mkdtemp()
        self.db_path = Path(self.temp_dir) / "test.db"
        self.employee_path = Path(self.temp_dir) / "employee.xlsx"
        self.export_dir = Path(self.temp_dir) / "exports"
        self.export_dir.mkdir()

        # Create employee file with Thai names
        wb = Workbook()
        ws = wb.active
        ws.append(["Legacy ID", "Full Name", "SL L1 Desc", "Position Desc"])
        ws.append(["TH001", "สมชาย ใจดี", "ไอที", "วิศวกร"])
        ws.append(["TH002", "สมหญิง รักเรียน", "ทรัพยากรบุคคล", "ผู้จัดการ"])
        ws.append(["MIX001", "John สมิธ", "IT", "Developer"])  # Mixed
        wb.save(self.employee_path)

    def tearDown(self):
        """Clean up."""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_thai_names_exported_correctly(self):
        """Test Thai employee names are preserved in export."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            # Scan Thai employee
            service.register_scan("TH001")

            result = service.export_scans()
            self.assertTrue(result["ok"])

            # Read export and verify Thai name
            wb = load_workbook(result["absolutePath"])
            ws = wb.active

            # Find the row with TH001
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == "TH001":
                    # Check name column contains Thai
                    row_data = list(row)
                    # Full Name should be in the row
                    self.assertTrue(any("สมชาย" in str(cell) for cell in row_data if cell))
                    break
            else:
                self.fail("TH001 not found in export")

            wb.close()
        finally:
            service.close()

    def test_mixed_thai_english_content(self):
        """Test mixed Thai/English content exports correctly."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            service.register_scan("MIX001")

            result = service.export_scans()
            self.assertTrue(result["ok"])

            wb = load_workbook(result["absolutePath"])
            ws = wb.active

            # Verify mixed content
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == "MIX001":
                    row_str = str(row)
                    self.assertIn("John", row_str)
                    self.assertIn("สมิธ", row_str)
                    break

            wb.close()
        finally:
            service.close()

    def test_thai_bu_names(self):
        """Test Thai BU names in export."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            service.register_scan("TH001")  # BU is "ไอที"

            result = service.export_scans()

            wb = load_workbook(result["absolutePath"])
            ws = wb.active

            # Find Thai BU
            found_thai_bu = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_str = str(row)
                if "ไอที" in row_str:
                    found_thai_bu = True
                    break

            self.assertTrue(found_thai_bu, "Thai BU name not found in export")
            wb.close()
        finally:
            service.close()


@unittest.skipUnless(PYQT6_AVAILABLE, "PyQt6 not available")
class TestExcelExportSpecialCharacters(unittest.TestCase):
    """Test special character handling in exports."""

    def setUp(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.db_path = Path(self.temp_dir) / "test.db"
        self.employee_path = Path(self.temp_dir) / "employee.xlsx"
        self.export_dir = Path(self.temp_dir) / "exports"
        self.export_dir.mkdir()

        # Create employee file
        wb = Workbook()
        ws = wb.active
        ws.append(["Legacy ID", "Full Name", "SL L1 Desc", "Position Desc"])
        ws.append(["TEST001", "O'Brien", "IT", "Engineer"])
        ws.append(["TEST002", 'John "Jack" Smith', "HR", "Manager"])
        ws.append(["TEST003", "José García", "Sales", "Rep"])
        wb.save(self.employee_path)

    def tearDown(self):
        """Clean up."""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_apostrophe_in_name(self):
        """Test apostrophe in employee name."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            service.register_scan("TEST001")
            result = service.export_scans()
            self.assertTrue(result["ok"])

            wb = load_workbook(result["absolutePath"])
            ws = wb.active

            found = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if "O'Brien" in str(row):
                    found = True
                    break

            self.assertTrue(found, "Apostrophe name not found")
            wb.close()
        finally:
            service.close()

    def test_quotes_in_name(self):
        """Test quotes in employee name."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            service.register_scan("TEST002")
            result = service.export_scans()
            self.assertTrue(result["ok"])

            wb = load_workbook(result["absolutePath"])
            ws = wb.active

            found = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if "Jack" in str(row):
                    found = True
                    break

            self.assertTrue(found, "Quote name not found")
            wb.close()
        finally:
            service.close()

    def test_accented_characters(self):
        """Test accented characters (José García)."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            service.register_scan("TEST003")
            result = service.export_scans()
            self.assertTrue(result["ok"])

            wb = load_workbook(result["absolutePath"])
            ws = wb.active

            found = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if "José" in str(row) or "García" in str(row):
                    found = True
                    break

            self.assertTrue(found, "Accented name not found")
            wb.close()
        finally:
            service.close()

    def test_special_badge_id(self):
        """Test special characters in badge ID are preserved."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            # Scan with special badge ID (unmatched employee)
            service.register_scan("BADGE-123_ABC")
            result = service.export_scans()

            wb = load_workbook(result["absolutePath"])
            ws = wb.active

            # Find the badge ID in export
            found = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == "BADGE-123_ABC":
                    found = True
                    break

            self.assertTrue(found, "Special badge ID not preserved")
            wb.close()
        finally:
            service.close()


@unittest.skipUnless(PYQT6_AVAILABLE, "PyQt6 not available")
class TestExcelExportEdgeCases(unittest.TestCase):
    """Test edge cases in export functionality."""

    def setUp(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.db_path = Path(self.temp_dir) / "test.db"
        self.employee_path = Path(self.temp_dir) / "employee.xlsx"
        self.export_dir = Path(self.temp_dir) / "exports"
        self.export_dir.mkdir()

        # Create employee file
        wb = Workbook()
        ws = wb.active
        ws.append(["Legacy ID", "Full Name", "SL L1 Desc", "Position Desc"])
        ws.append(["TEST001", "Test User", "IT", "Engineer"])
        wb.save(self.employee_path)

    def tearDown(self):
        """Clean up."""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_export_with_no_scans(self):
        """Test export with no scan data returns appropriate message."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            result = service.export_scans()

            self.assertFalse(result["ok"])
            self.assertTrue(result.get("noData", False))
            self.assertEqual(result["records"], 0)
        finally:
            service.close()

    def test_export_large_dataset(self):
        """Test export with large number of scans (1000+)."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            # Add many scans
            employee = service._employee_cache.get("TEST001")
            for i in range(1000):
                service._db.record_scan(f"BADGE{i:04d}", service.station_name, employee)

            result = service.export_scans()

            self.assertTrue(result["ok"])
            self.assertEqual(result["records"], 1000)

            # Verify file is valid
            wb = load_workbook(result["absolutePath"])
            ws = wb.active
            row_count = sum(1 for _ in ws.iter_rows(min_row=2))
            self.assertEqual(row_count, 1000)
            wb.close()
        finally:
            service.close()

    def test_export_unmatched_employee(self):
        """Test export handles unmatched badge correctly."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            # Scan unknown badge
            service.register_scan("UNKNOWN123")

            result = service.export_scans()
            self.assertTrue(result["ok"])

            wb = load_workbook(result["absolutePath"])
            ws = wb.active

            # Find the unmatched row
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == "UNKNOWN123":
                    # Matched (col 8) should be "No"
                    self.assertEqual(row[8], "No")
                    break

            wb.close()
        finally:
            service.close()

    def test_export_creates_directory_if_missing(self):
        """Test export creates export directory if it doesn't exist."""
        # Use a new non-existent directory
        new_export_dir = Path(self.temp_dir) / "new_exports"

        service = _create_service(self.db_path, self.employee_path, new_export_dir)

        try:
            service.register_scan("TEST001")
            result = service.export_scans()

            self.assertTrue(result["ok"])
            self.assertTrue(new_export_dir.exists())
        finally:
            service.close()


@unittest.skipUnless(PYQT6_AVAILABLE, "PyQt6 not available")
class TestExcelExportColumnWidth(unittest.TestCase):
    """Test column width adjustments."""

    def setUp(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.db_path = Path(self.temp_dir) / "test.db"
        self.employee_path = Path(self.temp_dir) / "employee.xlsx"
        self.export_dir = Path(self.temp_dir) / "exports"
        self.export_dir.mkdir()

        # Create employee with long name
        wb = Workbook()
        ws = wb.active
        ws.append(["Legacy ID", "Full Name", "SL L1 Desc", "Position Desc"])
        ws.append(["TEST001", "A" * 100, "Very Long Business Unit Name Here", "Position"])
        wb.save(self.employee_path)

    def tearDown(self):
        """Clean up."""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_column_width_reasonable(self):
        """Test column widths are set to reasonable values."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            service.register_scan("TEST001")
            result = service.export_scans()

            wb = load_workbook(result["absolutePath"])
            ws = wb.active

            # Check that columns have widths set (not default)
            for col_letter in ['A', 'B', 'C', 'D']:
                width = ws.column_dimensions[col_letter].width
                # Widths should be set (not None) and reasonable (< 60 per code)
                if width:
                    self.assertLessEqual(width, 62)  # Max is 60 + 2

            wb.close()
        finally:
            service.close()


@unittest.skipUnless(PYQT6_AVAILABLE, "PyQt6 not available")
class TestExcelExportMatchedColumn(unittest.TestCase):
    """Test the Matched Yes/No column."""

    def setUp(self):
        """Set up test environment."""
        self.temp_dir = tempfile.mkdtemp()
        self.db_path = Path(self.temp_dir) / "test.db"
        self.employee_path = Path(self.temp_dir) / "employee.xlsx"
        self.export_dir = Path(self.temp_dir) / "exports"
        self.export_dir.mkdir()

        wb = Workbook()
        ws = wb.active
        ws.append(["Legacy ID", "Full Name", "SL L1 Desc", "Position Desc"])
        ws.append(["TEST001", "Test User", "IT", "Engineer"])
        wb.save(self.employee_path)

    def tearDown(self):
        """Clean up."""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_matched_yes_for_known_employee(self):
        """Test Matched column shows Yes for known employee."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            service.register_scan("TEST001")
            result = service.export_scans()

            wb = load_workbook(result["absolutePath"])
            ws = wb.active

            # Find Matched column index
            headers = [cell.value for cell in ws[1]]
            matched_idx = headers.index("Matched")

            # Check data row
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == "TEST001":
                    self.assertEqual(row[matched_idx], "Yes")
                    break

            wb.close()
        finally:
            service.close()

    def test_matched_no_for_unknown_badge(self):
        """Test Matched column shows No for unknown badge."""
        service = _create_service(self.db_path, self.employee_path, self.export_dir)

        try:
            service.register_scan("UNKNOWN999")
            result = service.export_scans()

            wb = load_workbook(result["absolutePath"])
            ws = wb.active

            headers = [cell.value for cell in ws[1]]
            matched_idx = headers.index("Matched")

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == "UNKNOWN999":
                    self.assertEqual(row[matched_idx], "No")
                    break

            wb.close()
        finally:
            service.close()


def main():
    """Run tests with summary."""
    print("=" * 70)
    print("EXCEL EXPORT TESTS")
    print("=" * 70)
    print()

    # Create test suite
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()

    suite.addTests(loader.loadTestsFromTestCase(TestExcelExportBasic))
    suite.addTests(loader.loadTestsFromTestCase(TestExcelExportThaiCharacters))
    suite.addTests(loader.loadTestsFromTestCase(TestExcelExportSpecialCharacters))
    suite.addTests(loader.loadTestsFromTestCase(TestExcelExportEdgeCases))
    suite.addTests(loader.loadTestsFromTestCase(TestExcelExportColumnWidth))
    suite.addTests(loader.loadTestsFromTestCase(TestExcelExportMatchedColumn))

    # Run with verbosity
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)

    print()
    print("=" * 70)
    print(f"Tests run: {result.testsRun}")
    print(f"Failures: {len(result.failures)}")
    print(f"Errors: {len(result.errors)}")
    print("=" * 70)

    return 0 if result.wasSuccessful() else 1


if __name__ == "__main__":
    sys.exit(main())
