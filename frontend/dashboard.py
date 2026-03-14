"""Dashboard service for multi-station scan reports (Issue #27).

This module provides centralized reporting by:
- Querying Cloud API for multi-station scan data
- Reading local SQLite for employee count (registered headcount)

The dashboard shows:
- Total registered employees (from local SQLite)
- Total unique badges scanned (from Cloud API - all stations)
- Attendance rate (scanned / registered)
- Per-station breakdown
- Excel export functionality
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, TYPE_CHECKING

import requests

if TYPE_CHECKING:
    from database import DatabaseManager

logger = logging.getLogger(__name__)


class DashboardService:
    """Service for fetching multi-station dashboard data via Cloud API."""

    def __init__(
        self,
        db_manager: "DatabaseManager",
        api_url: str,
        api_key: str,
        export_directory: Optional[Path] = None,
    ) -> None:
        """Initialize the dashboard service.

        Args:
            db_manager: Local SQLite database manager (for employee count)
            api_url: Cloud API base URL
            api_key: Cloud API authentication key
            export_directory: Directory where Excel exports will be saved
        """
        self._db_manager = db_manager
        self._api_url = api_url.rstrip("/")
        self._api_key = api_key
        self._export_directory = export_directory or Path.cwd() / "exports"
        self._timeout = 15  # seconds
        self._employee_cache = None
        self._employee_cache_loaded = False

    def _get_employee_cache(self):
        """Get employee cache, loading once per session."""
        if not self._employee_cache_loaded:
            self._employee_cache = self._db_manager.load_employee_cache()
            self._employee_cache_loaded = True
            logger.debug(f"Dashboard: Loaded {len(self._employee_cache)} employees (cached)")
        return self._employee_cache

    def _get_headers(self) -> Dict[str, str]:
        """Get HTTP headers for API requests."""
        return {
            "Authorization": f"Bearer {self._api_key}",
            "Content-Type": "application/json",
        }

    def get_dashboard_data(self) -> Dict[str, Any]:
        """Fetch all dashboard data.

        Returns:
            Dictionary containing:
            - registered: int (employee count from local SQLite)
            - scanned: int (unique badges from cloud)
            - total_scans: int (total scans from cloud)
            - attendance_rate: float (percentage)
            - stations: list of station data
            - business_units: list of BU data with registered, scanned, and percentage
            - last_updated: str (ISO timestamp)
            - error: str (if any error occurred)
        """
        result = {
            "registered": 0,
            "scanned": 0,
            "total_scans": 0,
            "attendance_rate": 0.0,
            "stations": [],
            "business_units": [],
            "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "error": None,
        }

        # Get registered employee count from local SQLite
        try:
            result["registered"] = self._db_manager.count_employees()
            logger.debug(f"Dashboard: Local employee count = {result['registered']}")
        except Exception as e:
            logger.error(f"Dashboard: Failed to get employee count: {e}")
            result["error"] = f"Failed to get employee count: {e}"

        # Get cloud scan data from API
        cloud_bus = []  # BU data from cloud (all stations combined)
        try:
            response = requests.get(
                f"{self._api_url}/v1/dashboard/stats",
                headers=self._get_headers(),
                timeout=self._timeout,
            )

            if response.status_code == 200:
                data = response.json()
                result["total_scans"] = data.get("total_scans", 0)
                result["scanned"] = data.get("unique_badges", 0)
                cloud_bus = data.get("business_units", [])
                result["stations"] = sorted([
                    {
                        "name": s.get("name", "--"),
                        "scans": s.get("scans", 0),
                        "unique": s.get("unique", 0),
                        "last_scan": self._format_time(s.get("last_scan")),
                    }
                    for s in data.get("stations", [])
                ], key=lambda s: s["name"])
                logger.info(
                    f"Dashboard: total_scans={result['total_scans']}, "
                    f"unique_badges={result['scanned']}, stations={len(result['stations'])}"
                )
            elif response.status_code == 401:
                result["error"] = "Authentication failed - check API key"
                logger.error("Dashboard: API authentication failed")
            elif response.status_code == 503:
                result["error"] = "Cloud database unavailable"
                logger.error("Dashboard: Cloud database unavailable")
            else:
                result["error"] = f"API error: {response.status_code}"
                logger.error(f"Dashboard: API returned {response.status_code}")

        except requests.exceptions.ConnectionError:
            result["error"] = "Cannot connect to cloud API"
            logger.error("Dashboard: Connection error")
        except requests.exceptions.Timeout:
            result["error"] = "Cloud API timeout"
            logger.error("Dashboard: Request timeout")
        except Exception as e:
            result["error"] = f"API error: {e}"
            logger.error(f"Dashboard: Unexpected error: {e}")

        # Calculate attendance rate
        if result["registered"] > 0:
            result["attendance_rate"] = round(
                (result["scanned"] / result["registered"]) * 100, 1
            )

        # Get BU breakdown from cloud API (aggregates all stations)
        # Falls back to local SQLite if cloud data unavailable
        if cloud_bus:
            bu_list = []
            for bu in cloud_bus:
                registered = bu.get("registered", 0)
                scanned = bu.get("unique", 0)
                name = bu.get("name", "--")
                rate = round((scanned / registered) * 100, 1) if registered > 0 else 0.0
                bu_list.append({
                    "bu_name": name,
                    "registered": registered,
                    "scanned": scanned,
                    "attendance_rate": rate,
                })
            result["business_units"] = bu_list
            logger.info(f"Dashboard: BU breakdown from cloud API for {len(bu_list)} BUs")
        else:
            # Fallback: local SQLite (only this station's scans)
            try:
                bu_data = self._db_manager.get_scans_by_bu()
                bu_list = []
                for bu in bu_data:
                    registered = bu["registered"]
                    scanned = bu["scanned"]
                    rate = round((scanned / registered) * 100, 1) if registered > 0 else 0.0
                    bu_list.append({
                        "bu_name": bu["bu_name"],
                        "registered": registered,
                        "scanned": scanned,
                        "attendance_rate": rate,
                    })

                unmatched_count = self._db_manager.count_unmatched_scanned_badges()
                if unmatched_count > 0:
                    bu_list.append({
                        "bu_name": "(Unmatched)",
                        "registered": 0,
                        "scanned": unmatched_count,
                        "attendance_rate": 0.0,
                    })

                result["business_units"] = bu_list
                logger.info(f"Dashboard: BU breakdown from local DB for {len(bu_list)} BUs (cloud unavailable)")
            except Exception as e:
                logger.error(f"Dashboard: Failed to calculate BU breakdown: {e}")
                result["business_units"] = []

        return result

    def _format_time(self, iso_timestamp: Optional[str]) -> str:
        """Format ISO timestamp to time-only string."""
        if not iso_timestamp:
            return "--"
        try:
            dt = datetime.fromisoformat(iso_timestamp.replace("Z", "+00:00"))
            local_dt = dt.astimezone()  # Convert UTC to local machine timezone
            return local_dt.strftime("%H:%M:%S")
        except Exception:
            return "--"

    def _format_datetime(self, iso_timestamp: Optional[str]) -> str:
        """Format ISO timestamp to local date+time string for Excel export."""
        if not iso_timestamp:
            return "--"
        try:
            dt = datetime.fromisoformat(iso_timestamp.replace("Z", "+00:00"))
            local_dt = dt.astimezone()  # Convert UTC to local machine timezone
            return local_dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return iso_timestamp  # Return raw value as fallback

    def export_to_excel(self, dashboard_data: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """Export dashboard data to Excel file.

        Args:
            dashboard_data: Pre-fetched dashboard data to avoid redundant API calls.
                            If None, will fetch fresh data.

        Returns:
            Dictionary with:
            - ok: bool (success status)
            - message: str (success/error message)
            - file_path: str (absolute path to created file)
            - fileName: str (just the filename)
        """
        # Build export path with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Dashboard_Report_{timestamp}.xlsx"
        file_path = self._export_directory / filename

        # Ensure export directory exists
        self._export_directory.mkdir(parents=True, exist_ok=True)

        result = {
            "ok": False,
            "message": "",
            "file_path": str(file_path),
            "fileName": filename,
        }

        # Fetch export data from API
        try:
            response = requests.get(
                f"{self._api_url}/v1/dashboard/export",
                headers=self._get_headers(),
                timeout=60,  # Longer timeout for export
            )

            if response.status_code != 200:
                result["message"] = f"API error: {response.status_code}"
                return result

            data = response.json()
            scans = data.get("scans", [])

            if not scans:
                result["message"] = "No scan data to export"
                result["noData"] = True
                return result

        except requests.exceptions.ConnectionError:
            result["message"] = "Cannot connect to cloud API"
            return result
        except requests.exceptions.Timeout:
            result["message"] = "Cloud API timeout"
            return result
        except Exception as e:
            result["message"] = f"API error: {e}"
            return result

        # Generate Excel file
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            # Load employee cache for enriching scan data with employee details
            employee_cache = self._get_employee_cache()
            logger.debug(f"Dashboard export: Loaded {len(employee_cache)} employees for enrichment")

            # Enrich scans with employee details from local database
            enriched_scans = []
            for scan in scans:
                # Handle both dict and list formats from API
                if isinstance(scan, dict):
                    badge_id = scan.get("badge_id", "")
                    station = scan.get("station_name", "")
                    scanned_at = scan.get("scanned_at", "")
                    matched = scan.get("matched", False)
                else:
                    badge_id, station, scanned_at, matched = scan[0], scan[1], scan[2], scan[3]

                # Use legacy_id from cloud meta if available, otherwise try badge_id
                legacy_id = scan.get("legacy_id") if isinstance(scan, dict) else None
                lookup_key = legacy_id or badge_id
                employee = employee_cache.get(lookup_key)
                if employee:
                    full_name = employee.full_name
                    business_unit = employee.sl_l1_desc or "--"
                    position = employee.position_desc or "--"
                else:
                    full_name = "Unknown"
                    business_unit = "--"
                    position = "--"

                scan_source = scan.get("scan_source", "manual") if isinstance(scan, dict) else "manual"
                enriched_scans.append({
                    "Scan Value": badge_id,
                    "Legacy ID": employee.legacy_id if employee else (legacy_id or ""),
                    "Full Name": full_name,
                    "SL L1 Desc": business_unit,
                    "Position Desc": position,
                    "Email": employee.email if employee else "",
                    "Station": station,
                    "Scanned At": self._format_datetime(scanned_at),
                    "Matched": "Yes" if matched else "No",
                    "Scan Source": scan_source,
                })

            df = pd.DataFrame(enriched_scans)

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "All Scans"

            # Add header with styling
            header_fill = PatternFill(start_color="86bc25", end_color="86bc25", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            columns = ["Scan Value", "Legacy ID", "Full Name", "SL L1 Desc", "Position Desc", "Email", "Station", "Scanned At", "Matched", "Scan Source"]
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Add data rows
            for row_idx, row in enumerate(df.values, start=2):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws.column_dimensions[column].width = max_length + 2

            # Add summary sheet
            ws_summary = wb.create_sheet("Summary")
            if dashboard_data is None:
                dashboard_data = self.get_dashboard_data()

            ws_summary["A1"] = "Metric"
            ws_summary["B1"] = "Value"
            ws_summary["A1"].font = header_font
            ws_summary["A1"].fill = header_fill
            ws_summary["B1"].font = header_font
            ws_summary["B1"].fill = header_fill

            summary_data = [
                ("Registered Employees", dashboard_data["registered"]),
                ("Unique Badges Scanned", dashboard_data["scanned"]),
                ("Total Scans", dashboard_data["total_scans"]),
                ("Attendance Rate", f"{dashboard_data['attendance_rate']}%"),
                ("Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ]

            for row_idx, (label, value) in enumerate(summary_data, start=2):
                ws_summary.cell(row=row_idx, column=1, value=label)
                ws_summary.cell(row=row_idx, column=2, value=value)

            ws_summary.column_dimensions["A"].width = 25
            ws_summary.column_dimensions["B"].width = 20

            # Add station breakdown sheet
            if dashboard_data["stations"]:
                ws_stations = wb.create_sheet("By Station")
                station_headers = ["Station", "Total Scans", "Unique Badges", "Last Scan"]

                for col_idx, col_name in enumerate(station_headers, start=1):
                    cell = ws_stations.cell(row=1, column=col_idx, value=col_name)
                    cell.fill = header_fill
                    cell.font = header_font

                for row_idx, station in enumerate(dashboard_data["stations"], start=2):
                    ws_stations.cell(row=row_idx, column=1, value=station["name"])
                    ws_stations.cell(row=row_idx, column=2, value=station["scans"])
                    ws_stations.cell(row=row_idx, column=3, value=station["unique"])
                    ws_stations.cell(row=row_idx, column=4, value=station["last_scan"])

                for col in ws_stations.columns:
                    max_length = max(len(str(cell.value or "")) for cell in col)
                    ws_stations.column_dimensions[col[0].column_letter].width = max_length + 2

            # Add BU breakdown sheet — computed from enriched scan data
            # for consistency with "All Scans" sheet (fixes #52: stale cloud BU data)
            from collections import defaultdict
            bu_scanned = defaultdict(set)  # BU name → set of unique badge_ids
            for row in enriched_scans:
                badge_id = row["Scan Value"]
                bu = row["SL L1 Desc"]
                bu_key = bu if bu and bu != "--" else "(Unmatched)"
                bu_scanned[bu_key].add(badge_id)

            bu_registered = defaultdict(int)
            for emp in employee_cache.values():
                bu_registered[emp.sl_l1_desc or "(Unmatched)"] += 1

            all_bus = sorted(set(bu_scanned.keys()) | set(bu_registered.keys()))
            bu_export = []
            for bu_name in all_bus:
                if bu_name == "(Unmatched)":
                    continue  # append last
                registered = bu_registered.get(bu_name, 0)
                scanned = len(bu_scanned.get(bu_name, set()))
                rate = round((scanned / registered) * 100, 1) if registered > 0 else 0.0
                bu_export.append({"bu_name": bu_name, "registered": registered, "scanned": scanned, "attendance_rate": rate})
            if "(Unmatched)" in bu_scanned:
                bu_export.append({"bu_name": "(Unmatched)", "registered": 0,
                                  "scanned": len(bu_scanned["(Unmatched)"]), "attendance_rate": 0.0})

            if bu_export:
                ws_bu = wb.create_sheet("By Business Unit")
                bu_headers = ["Business Unit", "Registered", "Scanned", "Attendance %"]

                for col_idx, col_name in enumerate(bu_headers, start=1):
                    cell = ws_bu.cell(row=1, column=col_idx, value=col_name)
                    cell.fill = header_fill
                    cell.font = header_font

                for row_idx, bu in enumerate(bu_export, start=2):
                    ws_bu.cell(row=row_idx, column=1, value=bu["bu_name"])
                    ws_bu.cell(row=row_idx, column=2, value=bu["registered"])
                    ws_bu.cell(row=row_idx, column=3, value=bu["scanned"])
                    ws_bu.cell(row=row_idx, column=4, value=f"{bu['attendance_rate']}%")

                for col in ws_bu.columns:
                    max_length = max(len(str(cell.value or "")) for cell in col)
                    ws_bu.column_dimensions[col[0].column_letter].width = max_length + 2

            # Add "Not Yet Scanned" sheet - employees who haven't scanned
            scanned_badge_ids = {scan.get("badge_id") if isinstance(scan, dict) else scan[0] for scan in scans}
            not_scanned = [
                emp for emp in employee_cache.values()
                if emp.legacy_id not in scanned_badge_ids
            ]
            not_scanned.sort(key=lambda e: (e.sl_l1_desc or "", e.full_name or ""))

            ws_missing = wb.create_sheet("Not Yet Scanned")
            missing_headers = ["Legacy ID", "Full Name", "Email", "SL L1 Desc", "Position Desc"]
            for col_idx, col_name in enumerate(missing_headers, start=1):
                cell = ws_missing.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_idx, emp in enumerate(not_scanned, start=2):
                ws_missing.cell(row=row_idx, column=1, value=emp.legacy_id)
                ws_missing.cell(row=row_idx, column=2, value=emp.full_name)
                ws_missing.cell(row=row_idx, column=3, value=emp.email or "--")
                ws_missing.cell(row=row_idx, column=4, value=emp.sl_l1_desc or "--")
                ws_missing.cell(row=row_idx, column=5, value=emp.position_desc or "--")

            # Summary row
            summary_row = len(not_scanned) + 2
            ws_missing.cell(row=summary_row, column=1, value="Total Not Scanned:")
            ws_missing.cell(row=summary_row, column=1).font = Font(bold=True)
            ws_missing.cell(row=summary_row, column=2, value=len(not_scanned))
            ws_missing.cell(row=summary_row, column=2).font = Font(bold=True)

            for col in ws_missing.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                ws_missing.column_dimensions[col[0].column_letter].width = max_length + 2

            logger.info(f"Dashboard export: {len(not_scanned)} employees not yet scanned")

            # Save file
            wb.save(file_path)
            result["ok"] = True
            result["message"] = f"Exported {len(df)} scans to Excel"
            logger.info(f"Dashboard: Exported to {file_path}")

        except ImportError as e:
            result["message"] = f"Missing dependency: {e}. Run: pip install pandas openpyxl"
            logger.error(result["message"])
        except Exception as e:
            result["message"] = f"Export failed: {e}"
            logger.error(f"Dashboard export error: {e}")

        return result


__all__ = ["DashboardService"]
